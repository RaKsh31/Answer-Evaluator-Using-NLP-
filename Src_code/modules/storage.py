"""
modules/storage.py — Excel Results Storage Module
==================================================
Stores per-student evaluation results in an Excel workbook.
One sheet per subject. Columns:
  Student Name | Register Number | Q1 | Q2 | ... | Total | Percentage

Uses openpyxl for full formatting control.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour Palette ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # Dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # Mid blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")   # Light blue
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


# ── Save / Append Result ──────────────────────────────────────────────────────

def save_to_excel(filepath: str, subject: str, student_name: str,
                  reg_number: str, results: dict):
    """
    Append a student's evaluation result to the correct subject sheet.
    Creates the file or sheet if they don't exist.
    """
    # ── Load or create workbook ───────────────────────────────────────────────
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── Get or create the subject sheet ──────────────────────────────────────
    if subject in wb.sheetnames:
        ws = wb[subject]
    else:
        ws = wb.create_sheet(title=subject)
        _write_headers(ws, results)

    # ── Determine next row ────────────────────────────────────────────────────
    next_row = ws.max_row + 1

    # ── Build row data ────────────────────────────────────────────────────────
    q_labels = [q["label"] for q in results["questions"]]
    q_marks  = [q["marks"] for q in results["questions"]]

    row_data = (
        [student_name, reg_number]
        + q_marks
        + [results["total_marks"], f"{results['percentage']}%"]
    )

    # ── Write row with alternating fill ──────────────────────────────────────
    fill = ALT_ROW_FILL if next_row % 2 == 0 else None
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

    wb.save(filepath)


# ── Header Writer ─────────────────────────────────────────────────────────────

def _write_headers(ws, results: dict):
    """
    Write the header row for a new subject sheet.
    Columns: Student Name | Reg No | Q1 | Q2 | ... | Total | Percentage
    """
    q_labels = [q["label"] for q in results["questions"]]
    headers  = ["Student Name", "Register No."] + q_labels + ["Total", "Percentage"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Auto-width columns ────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)

    ws.row_dimensions[1].height = 30


# ── Subject List ──────────────────────────────────────────────────────────────

def get_subjects(filepath: str) -> list:
    """Return a list of sheet names (subjects) from the results workbook."""
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath, read_only=True)
    return wb.sheetnames


# ── Fetch Results for Subject ─────────────────────────────────────────────────

def get_results_for_subject(filepath: str, subject: str) -> dict:
    """
    Read all rows from a subject sheet and return as a list of dicts.
    Used by the /results/<subject> API endpoint.
    """
    if not os.path.exists(filepath):
        return {"columns": [], "rows": []}

    wb = load_workbook(filepath, read_only=True)
    if subject not in wb.sheetnames:
        return {"columns": [], "rows": []}

    ws = wb[subject]
    rows = list(ws.values)

    if not rows:
        return {"columns": [], "rows": []}

    columns = [str(c) for c in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append(dict(zip(columns, [str(v) if v is not None else "" for v in row])))

    return {"columns": columns, "rows": data_rows}
